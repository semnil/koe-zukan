"""
tests/test_build.py — build.py unit tests

Usage:
    python -m pytest tests/ -v
"""

import html as html_mod
import json
import os
import sys
from pathlib import Path
from unittest.mock import MagicMock, patch
from urllib.parse import urlparse

import openpyxl
import pytest

# Ensure scripts/ is importable
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))

import build


# ── _build_audio_ref ────────────────────────────────────

class TestBuildAudioRef:
    """Audio reference URL generation."""

    def test_bird_xeno_canto(self):
        url = build._build_audio_ref("B001", "Passer montanus", "pasmon1")
        assert url == "https://xeno-canto.org/species/Passer-montanus"

    def test_bird_no_scientific_name(self):
        assert build._build_audio_ref("B001", "", "pasmon1") == ""

    def test_bird_single_word_scientific_name(self):
        assert build._build_audio_ref("B001", "Passer", "pasmon1") == ""

    def test_mammal_macaulay(self):
        url = build._build_audio_ref("M001", "Canis lupus", "canlup1")
        assert "macaulaylibrary.org" in url
        assert "taxonCode=canlup1" in url

    def test_mammal_no_audio(self):
        """taxonCode in _NO_AUDIO list should return empty."""
        with patch.object(build, "_NO_AUDIO", {"canlup1"}):
            assert build._build_audio_ref("M001", "Canis lupus", "canlup1") == ""

    def test_mammal_empty_taxon_code(self):
        assert build._build_audio_ref("M001", "Canis lupus", "") == ""

    def test_mammal_none_taxon_code(self):
        """V004 fix: None taxon_code should not generate URL."""
        assert build._build_audio_ref("M001", "Canis lupus", None) == ""

    def test_insect_id_prefix(self):
        url = build._build_audio_ref("I001", "Gryllus bimaculatus", "grybim1")
        assert "macaulaylibrary.org" in url

    def test_amphibian_id_prefix(self):
        url = build._build_audio_ref("F001", "Rana japonica", "ranjap1")
        assert "macaulaylibrary.org" in url

    def test_initial_sample_prefix(self):
        """A-prefix uses Macaulay Library (only B-prefix uses xeno-canto)."""
        url = build._build_audio_ref("A001", "Felis catus", "felcat1")
        assert "macaulaylibrary.org" in url


# ── build_stats ─────────────────────────────────────────

class TestBuildStats:
    def _make_animals(self, ono_langs):
        return [{
            "onomatopoeia": [{"lang": l} for l in ono_langs],
        }]

    def test_basic_stats(self):
        animals = self._make_animals(["ja", "en", "ko"])
        stats = build.build_stats(animals)
        assert stats["total_species"] == 1
        assert stats["total_onomatopoeia"] == 3
        assert stats["language_count"] == 3
        assert sorted(stats["languages"]) == ["en", "ja", "ko"]

    def test_empty_animals(self):
        stats = build.build_stats([])
        assert stats["total_species"] == 0
        assert stats["total_onomatopoeia"] == 0
        assert stats["language_count"] == 0

    def test_duplicate_languages(self):
        animals = self._make_animals(["ja", "ja", "en"])
        stats = build.build_stats(animals)
        assert stats["language_count"] == 2

    def test_empty_lang_ignored(self):
        animals = self._make_animals(["ja", "", "en"])
        stats = build.build_stats(animals)
        assert stats["language_count"] == 2
        assert "" not in stats["languages"]


# ── generate_html (template placeholder replacement) ────

class TestIndexSeoMetadata:
    """The built top page must carry canonical + WebSite JSON-LD so duplicate
    URL variants (?lang=, ?id=, ?q=) collapse to one canonical entry and the
    site qualifies for Google's Sitelinks Search Box."""

    @pytest.fixture(scope="class")
    def html(self):
        import subprocess
        root = Path(__file__).resolve().parent.parent
        if not (root / "dist" / "index.html").exists():
            subprocess.run(
                ["python3", "scripts/build.py"],
                cwd=root, capture_output=True, check=True,
            )
        return (root / "dist" / "index.html").read_text(encoding="utf-8")

    def test_canonical_present(self, html):
        assert f'<link rel="canonical" href="{build.SITE_URL}/">' in html

    def test_website_json_ld_block_parses(self, html):
        import re
        # First JSON-LD block on the top page must be the WebSite descriptor.
        m = re.search(
            r'<script type="application/ld\+json">\s*(\{.*?\})\s*</script>',
            html, re.DOTALL,
        )
        assert m, "no JSON-LD on index"
        data = json.loads(m.group(1))
        assert data["@type"] == "WebSite"
        assert data["url"] == f"{build.SITE_URL}/"
        assert data["inLanguage"] == ["ja", "en", "ko", "zh"]

    def test_website_json_ld_search_action(self, html):
        import re
        m = re.search(
            r'<script type="application/ld\+json">\s*(\{.*?\})\s*</script>',
            html, re.DOTALL,
        )
        data = json.loads(m.group(1))
        action = data.get("potentialAction") or {}
        assert action.get("@type") == "SearchAction"
        # urlTemplate is the entry point Google uses to invoke the search box.
        target = action.get("target") or {}
        assert target.get("urlTemplate") == f"{build.SITE_URL}/?q={{search_term_string}}"
        assert action.get("query-input") == "required name=search_term_string"


class TestGenerateHtml:
    def test_placeholder_replacement(self, tmp_path):
        template = tmp_path / "template.html"
        template.write_text(
            "<p>{{SITE_URL}}</p>"
            "<p>{{SPECIES_COUNT}} species</p>"
            "<p>{{LANGUAGE_COUNT}} langs</p>"
            "<p>{{ONOMATOPOEIA_COUNT}} onomatopoeia</p>",
            encoding="utf-8",
        )
        output = tmp_path / "index.html"
        animals = [
            {"onomatopoeia": [{"lang": "ja"}, {"lang": "en"}]},
            {"onomatopoeia": [{"lang": "ja"}]},
        ]
        build.generate_html(animals, template, output)
        html = output.read_text(encoding="utf-8")
        assert build.SITE_URL in html
        assert "2 species" in html
        assert "2 langs" in html
        assert "3 onomatopoeia" in html

    def test_missing_template_no_fallback(self, tmp_path):
        """No template and no fallback should print warning, not crash."""
        output = tmp_path / "index.html"
        build.generate_html([], tmp_path / "nonexistent.html", output)
        assert not output.exists()


# ── generate_sitemap ────────────────────────────────────

class TestGenerateSitemap:
    def test_sitemap_structure(self, tmp_path):
        animals = [{"id": "B001"}, {"id": "M001"}]
        out = tmp_path / "sitemap.xml"
        build.generate_sitemap(animals, out)
        xml = out.read_text(encoding="utf-8")
        assert '<?xml version="1.0"' in xml
        assert f"{build.SITE_URL}/" in xml
        assert "/species/" not in xml

    def test_sitemap_top_page_only(self, tmp_path):
        animals = [{"id": f"X{i:03d}"} for i in range(5)]
        out = tmp_path / "sitemap.xml"
        build.generate_sitemap(animals, out)
        xml = out.read_text(encoding="utf-8")
        assert xml.count("<url>") == 1  # top page only


# ── generate_manifest ──────────────────────────────────

class TestGenerateManifest:
    def test_manifest_valid_json(self, tmp_path):
        out = tmp_path / "manifest.json"
        build.generate_manifest(out)
        data = json.loads(out.read_text(encoding="utf-8"))
        assert data["name"] == "動物の鳴き声図鑑"
        assert data["display"] == "standalone"
        assert data["theme_color"] == "#047857"
        assert any(icon["src"] == "/favicon.svg" for icon in data["icons"])


# ── generate_sw ─────────────────────────────────────────

class TestGenerateServiceWorker:
    def test_sw_contains_cache_name(self, tmp_path):
        out = tmp_path / "sw.js"
        build.generate_sw([{"id": "B001"}], out)
        js = out.read_text(encoding="utf-8")
        assert "CACHE_NAME" in js
        assert "koe-zukan-v" in js

    def test_sw_contains_required_urls(self, tmp_path):
        out = tmp_path / "sw.js"
        build.generate_sw([], out)
        js = out.read_text(encoding="utf-8")
        for url in ["/", "/animals.json", "/regions.json", "/favicon.svg", "/manifest.json"]:
            assert url in js

    def test_sw_event_handlers(self, tmp_path):
        out = tmp_path / "sw.js"
        build.generate_sw([], out)
        js = out.read_text(encoding="utf-8")
        assert '"install"' in js
        assert '"activate"' in js
        assert '"fetch"' in js


# ── generate_species_pages ──────────────────────────────

class TestGenerateSpeciesPages:
    def _make_animal(self, aid="B001"):
        return {
            "id": aid,
            "nameJA": "スズメ",
            "nameEN": "Eurasian Tree Sparrow",
            "scientificName": "Passer montanus",
            "altJA": "すずめ",
            "altEN": "",
            "class": "鳥綱",
            "order": "スズメ目",
            "family": "スズメ科",
            "voiceMethod": "鳴管",
            "conservation": "LC",
            "habitat": "ユーラシア",
            "note": "",
            "imageRef": "https://commons.wikimedia.org/wiki/Category:Passer_montanus",
            "audioRef": "https://xeno-canto.org/species/Passer-montanus",
            "onomatopoeiaJA": "チュンチュン",
            "onomatopoeia": [
                {"lang": "ja", "onomatopoeia": "チュンチュン", "scene": "さえずり", "note": ""},
                {"lang": "en", "onomatopoeia": "Chirp chirp", "scene": "", "note": ""},
            ],
            "regions": [{"id": "R01", "nameJA": "日本", "nameEN": "Japan"}],
        }

    def test_species_page_generated(self, tmp_path):
        animal = self._make_animal()
        count = build.generate_species_pages([animal], tmp_path)
        assert count == 1
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        assert "スズメ" in page
        assert "Eurasian Tree Sparrow" in page
        assert "Passer montanus" in page

    def test_species_page_ogp_meta(self, tmp_path):
        animal = self._make_animal()
        build.generate_species_pages([animal], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        assert f'{build.SITE_URL}/species/B001/ogp.png' in page
        assert f'{build.SITE_URL}/species/B001/' in page

    def test_species_page_canonical_url(self, tmp_path):
        animal = self._make_animal()
        build.generate_species_pages([animal], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        assert f'rel="canonical" href="{build.SITE_URL}/species/B001/"' in page

    def test_species_page_json_ld(self, tmp_path):
        animal = self._make_animal()
        build.generate_species_pages([animal], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        assert '"@type": "Article"' in page
        assert '"@context": "https://schema.org"' in page

    def test_species_page_onomatopoeia_section(self, tmp_path):
        animal = self._make_animal()
        build.generate_species_pages([animal], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        assert "チュンチュン" in page
        assert "Chirp chirp" in page

    def test_species_page_share_buttons(self, tmp_path):
        animal = self._make_animal()
        build.generate_species_pages([animal], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        assert "twitter.com/intent/tweet" in page
        assert "facebook.com/sharer" in page
        assert "line.me/lineit/share" in page
        assert "copyShareUrl" in page

    def test_species_page_conservation_label(self, tmp_path):
        animal = self._make_animal()
        build.generate_species_pages([animal], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        assert "LC (低懸念)" in page

    def test_species_page_external_links(self, tmp_path):
        animal = self._make_animal()
        build.generate_species_pages([animal], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        assert "Wikimedia Commons" in page
        assert "xeno-canto" in page
        assert "Wikipedia (JA)" in page
        assert "Wikipedia (EN)" in page

    def test_species_page_html_escaping(self, tmp_path):
        """V001: Special chars in data fields must be HTML-escaped."""
        animal = self._make_animal()
        animal["note"] = 'Test <b>"note"</b> & more'
        build.generate_species_pages([animal], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        assert "Test &lt;b&gt;&quot;note&quot;&lt;/b&gt; &amp; more" in page
        assert '<b>"note"</b>' not in page

    def test_species_page_no_note(self, tmp_path):
        animal = self._make_animal()
        animal["note"] = ""
        build.generate_species_pages([animal], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        # 備考行 (detail-row + label) は描画されない。日本語ラベル文字列は i18n
        # 辞書にも含まれるため "備考" 単体での検査ではなく構造マーカで判定する。
        assert 'data-i18n="note"' not in page

    def test_species_page_with_note(self, tmp_path):
        animal = self._make_animal()
        animal["note"] = "テスト備考"
        build.generate_species_pages([animal], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        assert 'data-i18n="note"' in page
        assert "テスト備考" in page

    def test_species_page_no_voice_method(self, tmp_path):
        animal = self._make_animal()
        animal["voiceMethod"] = ""
        build.generate_species_pages([animal], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        assert "—" in page

    def test_species_page_back_link(self, tmp_path):
        animal = self._make_animal()
        build.generate_species_pages([animal], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        assert '/?id=B001' in page

    def test_species_page_related_section_present(self, tmp_path):
        """Related species section must appear when siblings exist."""
        sibling = self._make_animal("B002")
        sibling["nameJA"] = "ニワトリ"
        sibling["onomatopoeiaJA"] = "コケコッコ"
        animal = self._make_animal("B001")
        build.generate_species_pages([animal, sibling], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        assert "同じ仲間の動物" in page
        assert "/species/B002/" in page
        assert "ニワトリ" in page
        assert "コケコッコ" in page

    def test_species_page_related_excludes_self(self, tmp_path):
        animal = self._make_animal("B001")
        sibling = self._make_animal("B002")
        sibling["nameJA"] = "ニワトリ"
        sibling["onomatopoeiaJA"] = "コケコッコ"
        build.generate_species_pages([animal, sibling], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        # The related-item link must not point to self
        assert 'class="related-item" href="/species/B001/"' not in page

    def test_species_page_related_no_ono_renders(self, tmp_path):
        """Related item with empty onomatopoeiaJA must not crash."""
        sibling = self._make_animal("B002")
        sibling["nameJA"] = "ニワトリ"
        sibling["onomatopoeiaJA"] = ""
        animal = self._make_animal("B001")
        build.generate_species_pages([animal, sibling], tmp_path)
        page = (tmp_path / "species" / "B001" / "index.html").read_text(encoding="utf-8")
        assert "ニワトリ" in page


# ── _build_related_species ──────────────────────────────

class TestBuildRelatedSpecies:
    def _a(self, aid, family, order, ono="", class_="鳥綱", phylum="脊索動物門"):
        return {"id": aid, "family": family, "order": order, "class": class_, "phylum": phylum, "onomatopoeiaJA": ono}

    def test_same_family_returned(self):
        animals = [
            self._a("B001", "スズメ科", "スズメ目"),
            self._a("B002", "スズメ科", "スズメ目"),
            self._a("M001", "ネコ科", "食肉目", class_="哺乳綱", phylum="節足動物門"),
        ]
        related = build._build_related_species(animals[0], animals)
        ids = [r["id"] for r in related]
        assert "B002" in ids
        assert "M001" not in ids

    def test_self_excluded(self):
        animals = [self._a("B001", "スズメ科", "スズメ目")]
        related = build._build_related_species(animals[0], animals)
        assert not any(r["id"] == "B001" for r in related)

    def test_id_order_sorted(self):
        animals = [
            self._a("B003", "スズメ科", "スズメ目"),
            self._a("B001", "スズメ科", "スズメ目"),
            self._a("B002", "スズメ科", "スズメ目"),
        ]
        related = build._build_related_species(animals[0], animals)
        assert [r["id"] for r in related] == ["B001", "B002"]

    def test_max_four_returned(self):
        animals = [self._a(f"B{i:03d}", "スズメ科", "スズメ目") for i in range(10)]
        related = build._build_related_species(animals[0], animals)
        assert len(related) <= 4

    def test_order_fallback_when_family_alone(self):
        animals = [
            self._a("B001", "ユニーク科", "スズメ目"),
            self._a("B002", "別の科", "スズメ目"),
            self._a("M001", "ネコ科", "食肉目", class_="哺乳綱", phylum="節足動物門"),
        ]
        related = build._build_related_species(animals[0], animals)
        ids = [r["id"] for r in related]
        assert "B002" in ids
        assert "M001" not in ids

    def test_fallback_fills_up_to_four(self):
        # 1 same-family sibling + 5 same-order siblings → total 4
        animals = (
            [self._a("B001", "ユニーク科", "スズメ目")]
            + [self._a(f"B{i:03d}", "スズメ科", "スズメ目") for i in range(2, 8)]
        )
        related = build._build_related_species(animals[0], animals)
        assert len(related) == 4

    def test_empty_family_uses_order_only(self):
        animals = [
            self._a("B001", "", "スズメ目"),
            self._a("B002", "スズメ科", "スズメ目"),
        ]
        related = build._build_related_species(animals[0], animals)
        assert any(r["id"] == "B002" for r in related)

    def test_class_fallback_when_order_alone(self):
        animals = [
            self._a("B001", "ユニーク科", "ユニーク目", class_="鳥綱"),
            self._a("B002", "スズメ科", "スズメ目", class_="鳥綱"),
            self._a("M001", "ネコ科", "食肉目", class_="哺乳綱", phylum="節足動物門"),
        ]
        related = build._build_related_species(animals[0], animals)
        ids = [r["id"] for r in related]
        assert "B002" in ids
        assert "M001" not in ids

    def test_phylum_fallback_when_class_alone(self):
        animals = [
            self._a("V001", "ユニーク科", "ユニーク目", class_="ユニーク綱", phylum="軟体動物門"),
            self._a("V002", "別の科", "別の目", class_="別の綱", phylum="軟体動物門"),
            self._a("M001", "ネコ科", "食肉目", class_="哺乳綱", phylum="脊索動物門"),
        ]
        related = build._build_related_species(animals[0], animals)
        ids = [r["id"] for r in related]
        assert "V002" in ids
        assert "M001" not in ids

    def test_no_family_no_order_no_class_no_phylum_returns_empty(self):
        animals = [
            self._a("B001", "", "", class_="", phylum=""),
            self._a("B002", "スズメ科", "スズメ目", class_="鳥綱"),
        ]
        related = build._build_related_species(animals[0], animals)
        assert related == []


# ── _build_related_html ──────────────────────────────────

class TestBuildRelatedHtml:
    def _a(self, aid, name, ono=""):
        return {"id": aid, "nameJA": name, "onomatopoeiaJA": ono}

    def test_empty_list_returns_empty_string(self):
        assert build._build_related_html([], html_mod.escape) == ""

    def test_contains_heading(self):
        html = build._build_related_html([self._a("B001", "スズメ", "チュン")], html_mod.escape)
        assert "同じ仲間の動物" in html

    def test_contains_name_and_ono(self):
        html = build._build_related_html([self._a("B001", "スズメ", "チュン")], html_mod.escape)
        assert "スズメ" in html
        assert "チュン" in html

    def test_link_href_format(self):
        html = build._build_related_html([self._a("B001", "スズメ")], html_mod.escape)
        assert 'href="/species/B001/"' in html

    def test_no_ono_span_when_empty(self):
        html = build._build_related_html([self._a("B001", "スズメ", "")], html_mod.escape)
        assert "related-ono" not in html

    def test_html_escaping_in_name(self):
        html = build._build_related_html(
            [self._a("X001", '<script>alert(1)</script>', "テスト")], html_mod.escape
        )
        assert "<script>" not in html
        assert "&lt;script&gt;" in html


# ── _parse_svg_points ───────────────────────────────────

class TestParseSvgPoints:
    def test_basic_path(self, tmp_path):
        svg = tmp_path / "test.svg"
        svg.write_text('<svg><path d="M 10 20 L 30 40 50 60"/></svg>')
        points = build._parse_svg_points(svg)
        assert (10.0, 20.0) in points
        assert (30.0, 40.0) in points
        assert (50.0, 60.0) in points

    def test_no_path(self, tmp_path):
        svg = tmp_path / "test.svg"
        svg.write_text('<svg><circle cx="10" cy="10" r="5"/></svg>')
        points = build._parse_svg_points(svg)
        assert points == []


# ── _find_cjk_fonts ─────────────────────────────────────

class TestFindCjkFonts:
    def test_returns_all_keys(self):
        fonts = build._find_cjk_fonts()
        assert "ja" in fonts
        assert "ko" in fonts
        assert "zh" in fonts

    def test_ko_zh_fallback_to_ja(self):
        """If ko/zh font not found, should fall back to ja font."""
        fonts = build._find_cjk_fonts()
        # On any platform, ko and zh should never be None if ja is found
        if fonts["ja"]:
            assert fonts["ko"] is not None
            assert fonts["zh"] is not None


# ── CNAME generation ────────────────────────────────────

class TestCnameGeneration:
    def test_cname_from_site_url(self):
        domain = urlparse(build.SITE_URL).netloc
        assert domain == "koe-zukan.semnil.com"


# ── CONSERVATION_JA ─────────────────────────────────────

class TestConservationLabels:
    def test_all_iucn_categories_present(self):
        expected = {"LC", "NT", "VU", "EN", "CR", "DD", "NE", "EW", "EX"}
        assert set(build.CONSERVATION_JA.keys()) == expected

    def test_labels_are_non_empty(self):
        for key, val in build.CONSERVATION_JA.items():
            assert val, f"Empty label for {key}"


# ── LANG_LABELS ─────────────────────────────────────────

class TestLangLabels:
    def test_four_languages(self):
        assert set(build.LANG_LABELS.keys()) == {"ja", "en", "ko", "zh"}

    def test_labels_non_empty(self):
        for key, val in build.LANG_LABELS.items():
            assert val, f"Empty label for {key}"


# ── no-audio.json integrity ─────────────────────────────

class TestNoAudioJson:
    """Data file integrity: no-audio.json must have unique taxonCodes."""

    def test_no_audio_entries_are_unique(self):
        # Guards against duplicate entries slipping in via check_audio.py runs.
        # Duplicates waste memory and signal stale data.
        if not build.NO_AUDIO_FILE.exists():
            pytest.skip("no-audio.json not present")
        with open(build.NO_AUDIO_FILE, "r", encoding="utf-8") as f:
            items = json.load(f)
        assert len(items) == len(set(items)), (
            f"Duplicates in no-audio.json: "
            f"{[x for x in items if items.count(x) > 1]}"
        )

    def test_no_audio_entries_are_strings(self):
        if not build.NO_AUDIO_FILE.exists():
            pytest.skip("no-audio.json not present")
        with open(build.NO_AUDIO_FILE, "r", encoding="utf-8") as f:
            items = json.load(f)
        for it in items:
            assert isinstance(it, str) and it, f"Invalid taxonCode entry: {it!r}"


# ── JSON-LD injection safety ───────────────────────────

class TestJsonLdInjectionSafety:
    """JSON-LD script block must not be breakable by field values containing `</script>`.

    Python's `json.dumps` escapes `"` and `\\` but NOT `<`, `>`, `&`, so a field
    value like `</script><script>alert(1)</script>` would split the JSON-LD block
    and allow arbitrary script injection. This is a latent XSS; Excel data is
    trusted but the code contract should defend against it.
    """

    def _make_animal(self, nameJA="", nameEN="", scientificName="Sci"):
        return {
            "id": "X001",
            "nameJA": nameJA or "テスト",
            "nameEN": nameEN,
            "scientificName": scientificName,
            "altJA": "", "altEN": "",
            "class": "test", "order": "test", "family": "test",
            "voiceMethod": "", "conservation": "",
            "habitat": "", "note": "",
            "imageRef": "", "audioRef": "",
            "onomatopoeiaJA": "",
            "onomatopoeia": [],
            "regions": [],
        }

    def _json_ld_range(self, html):
        """Return (start, end_of_opening_tag) for the JSON-LD script block.

        `end_of_opening_tag` is the character position right after the opening
        `<script type="application/ld+json">`. The body extends from there
        until the NEXT `</script>` in the document. An injection succeeds when
        that `</script>` appears BEFORE the legitimate one.
        """
        import re
        open_m = re.search(
            r'<script type="application/ld\+json">', html
        )
        if not open_m:
            return None
        body_start = open_m.end()
        # Find the FIRST closing tag after body_start
        close_idx = html.find("</script>", body_start)
        return (body_start, close_idx)

    def test_json_ld_resists_closing_script_tag_in_nameEN(self, tmp_path):
        animal = self._make_animal(
            nameJA="テスト",
            nameEN='Evil</script><script>alert("xss")</script>',
        )
        build.generate_species_pages([animal], tmp_path)
        html = (tmp_path / "species" / "X001" / "index.html").read_text(
            encoding="utf-8",
        )
        # Expose the defect: the body BETWEEN the JSON-LD opening tag and the
        # first </script> must parse as valid JSON. If the attacker's </script>
        # closes the block prematurely, the body is truncated JSON.
        rng = self._json_ld_range(html)
        assert rng is not None
        body = html[rng[0]:rng[1]].strip()
        import json as _json
        try:
            _json.loads(body)
        except _json.JSONDecodeError as e:
            raise AssertionError(
                f"JSON-LD block truncated by injected </script>: {e}\n"
                f"Body starts with: {body[:200]!r}"
            )

    def test_json_ld_resists_closing_script_tag_in_nameJA(self, tmp_path):
        animal = self._make_animal(
            nameJA='悪意</script><script>alert(1)</script>',
            nameEN="Test",
        )
        build.generate_species_pages([animal], tmp_path)
        html = (tmp_path / "species" / "X001" / "index.html").read_text(
            encoding="utf-8",
        )
        rng = self._json_ld_range(html)
        assert rng is not None
        body = html[rng[0]:rng[1]].strip()
        import json as _json
        try:
            _json.loads(body)
        except _json.JSONDecodeError as e:
            raise AssertionError(
                f"JSON-LD block truncated by injected </script>: {e}"
            )


# ── Build integration (full pipeline) ───────────────────

class TestBuildIntegration:
    """Run the full pipeline against the real Excel and assert key invariants on dist/."""

    @pytest.fixture(scope="class")
    def dist(self):
        import subprocess
        root = Path(__file__).resolve().parent.parent
        result = subprocess.run(
            ["python3", "scripts/build.py"],
            cwd=root, capture_output=True, text=True,
        )
        assert result.returncode == 0, (
            f"build failed: {result.stderr[-500:]}"
        )
        return root / "dist"

    def test_sitemap_has_no_species_urls(self, dist):
        import re
        sitemap = (dist / "sitemap.xml").read_text(encoding="utf-8")
        sm_ids = set(re.findall(r"/species/([^/]+)/", sitemap))
        assert not sm_ids, f"sitemap should not contain species URLs: {sm_ids}"

    def test_no_unreplaced_placeholders_in_dist(self, dist):
        import re
        pattern = re.compile(r"\{\{[A-Z_]+\}\}")
        offenders = []
        for html in dist.rglob("*.html"):
            text = html.read_text(encoding="utf-8")
            found = pattern.findall(text)
            if found:
                offenders.append((html.relative_to(dist), found))
        assert not offenders, f"Unreplaced placeholders: {offenders[:5]}"

    def test_all_json_ld_blocks_parse(self, dist):
        import re
        bad = []
        for page in (dist / "species").rglob("index.html"):
            html = page.read_text(encoding="utf-8")
            m = re.search(
                r'<script type="application/ld\+json">\s*(\{.*?\})\s*</script>',
                html, re.DOTALL,
            )
            if not m:
                bad.append((page.parent.name, "missing"))
                continue
            try:
                data = json.loads(m.group(1))
                if data.get("@type") != "Article":
                    bad.append((page.parent.name, f"type={data.get('@type')}"))
            except json.JSONDecodeError as e:
                bad.append((page.parent.name, str(e)))
        assert not bad, f"JSON-LD issues: {bad[:5]}"

    def test_all_external_links_have_rel_noopener(self, dist):
        import re
        offenders = []
        for page in (dist / "species").rglob("index.html"):
            html = page.read_text(encoding="utf-8")
            for tag in re.findall(r'<a[^>]+target="_blank"[^>]*>', html):
                if "noopener" not in tag:
                    offenders.append((page.parent.name, tag[:80]))
                    break
        assert not offenders, f"External links missing rel=noopener: {offenders[:3]}"

    def test_no_audio_taxon_codes_not_in_any_audio_url(self, dist):
        with open(dist / "animals.json", encoding="utf-8") as f:
            animals = json.load(f)
        if not build.NO_AUDIO_FILE.exists():
            pytest.skip("no-audio.json absent")
        with open(build.NO_AUDIO_FILE, encoding="utf-8") as f:
            no_audio = set(json.load(f))
        # Only non-bird audio URLs can contain taxonCodes
        violations = []
        for a in animals:
            if a["id"].startswith("B"):
                continue
            for tc in no_audio:
                if tc in a.get("audioRef", ""):
                    violations.append((a["id"], tc))
        assert not violations, f"no-audio taxonCodes appearing in audioRef: {violations[:3]}"

    def test_species_ogp_images_nonempty(self, dist):
        zero_files = [
            p.parent.name
            for p in (dist / "species").rglob("ogp.png")
            if p.stat().st_size == 0
        ]
        assert not zero_files, f"Zero-byte OGP: {zero_files[:5]}"

    def test_sitemap_contains_top_page_only(self, dist):
        import re
        sitemap = (dist / "sitemap.xml").read_text(encoding="utf-8")
        urls = re.findall(r"<loc>([^<]+)</loc>", sitemap)
        assert len(urls) == 1, f"sitemap should contain only top page, got {len(urls)} URLs"
        assert urls[0] == f"{build.SITE_URL}/"

    def test_animals_json_ids_unique(self, dist):
        with open(dist / "animals.json", encoding="utf-8") as f:
            animals = json.load(f)
        ids = [a["id"] for a in animals]
        assert len(ids) == len(set(ids))

    def test_animals_json_id_format(self, dist):
        import re
        with open(dist / "animals.json", encoding="utf-8") as f:
            animals = json.load(f)
        bad = [a["id"] for a in animals if not re.match(r"^[A-Z][0-9]+$", a["id"])]
        assert not bad, f"IDs not matching /^[A-Z][0-9]+$/: {bad}"

    def test_all_audio_refs_are_https_urls(self, dist):
        with open(dist / "animals.json", encoding="utf-8") as f:
            animals = json.load(f)
        invalid = [a["id"] for a in animals if a.get("audioRef") and not a["audioRef"].startswith("https://")]
        assert not invalid, f"audioRef not https: {invalid[:3]}"

    def test_image_refs_are_https_urls(self, dist):
        with open(dist / "animals.json", encoding="utf-8") as f:
            animals = json.load(f)
        invalid = [a["id"] for a in animals if a.get("imageRef") and not a["imageRef"].startswith("https://")]
        assert not invalid, f"imageRef not https: {invalid[:3]}"

    def test_sw_cache_version_changes_with_species_count(self, dist):
        # Version is a function of date + species count; two consecutive runs on the
        # same day produce identical cache names, giving deterministic offline behavior.
        sw = (dist / "sw.js").read_text(encoding="utf-8")
        with open(dist / "animals.json", encoding="utf-8") as f:
            animals = json.load(f)
        assert f"-{len(animals)}" in sw

    def test_built_sw_is_syntactically_valid(self, dist):
        # Real dist/sw.js must parse. A syntax error silently breaks PWA
        # offline caching for every visitor.
        import shutil, subprocess
        if not shutil.which("node"):
            pytest.skip("node not available")
        result = subprocess.run(
            ["node", "--check", str(dist / "sw.js")],
            capture_output=True, text=True,
        )
        assert result.returncode == 0, (
            f"dist/sw.js syntax error: {result.stderr}"
        )

    def test_built_sitemap_is_well_formed_xml(self, dist):
        import xml.etree.ElementTree as ET
        ET.parse(dist / "sitemap.xml")  # raises on malformed XML

    def test_all_species_pages_have_related_section(self, dist):
        missing = [
            p.parent.name
            for p in (dist / "species").rglob("index.html")
            if "同じ仲間の動物" not in p.read_text(encoding="utf-8")
        ]
        assert not missing, f"Missing related section: {missing[:5]}"

    def test_related_links_point_to_valid_ids(self, dist):
        import re
        with open(dist / "animals.json", encoding="utf-8") as f:
            valid_ids = {a["id"] for a in json.load(f)}
        bad = []
        for page in (dist / "species").rglob("index.html"):
            html = page.read_text(encoding="utf-8")
            for mid in re.findall(r'class="related-item"[^>]*href="/species/([^/]+)/"', html):
                if mid not in valid_ids:
                    bad.append((page.parent.name, mid))
        assert not bad, f"Related links to invalid IDs: {bad[:5]}"

    def test_index_html_has_species_detail_link(self, dist):
        html = (dist / "index.html").read_text(encoding="utf-8")
        assert "species-detail-link" in html
        assert "詳細ページで見る" in html


# ── Species page i18n payload ───────────────────────────


class TestSpeciesPageI18nPayload:
    """The species page embeds a SPECIES_DATA JSON blob and language-tagged DOM
    markers so the client script can swap labels per display language. These
    tests document the contract between build.py and templates/species.html.
    """

    def _make_animal(self, aid="B001"):
        return {
            "id": aid,
            "nameJA": "スズメ",
            "nameEN": "Eurasian Tree Sparrow",
            "scientificName": "Passer montanus",
            "altJA": "すずめ",
            "altEN": "",
            "class": "鳥綱",
            "classEN": "Birds",
            "order": "スズメ目",
            "orderEN": "Passeriformes",
            "family": "スズメ科",
            "familyEN": "Passeridae",
            "voiceMethod": "鳴管",
            "conservation": "LC",
            "habitat": "ユーラシア",
            "note": "",
            "imageRef": "https://commons.wikimedia.org/wiki/Category:Passer_montanus",
            "audioRef": "https://xeno-canto.org/species/Passer-montanus",
            "onomatopoeiaJA": "チュンチュン",
            "hasVoice": "あり",
            "onomatopoeia": [
                {"lang": "ja", "onomatopoeia": "チュンチュン", "scene": "さえずり", "note": ""},
                {"lang": "en", "onomatopoeia": "Chirp chirp", "scene": "", "note": ""},
                {"lang": "ko", "onomatopoeia": "짹짹", "scene": "", "note": ""},
                {"lang": "zh", "onomatopoeia": "啾啾", "scene": "", "note": ""},
            ],
            "regions": [{"id": "R01", "nameJA": "日本", "nameEN": "Japan"}],
        }

    def _read_page(self, tmp_path, aid="B001"):
        animal = self._make_animal(aid)
        build.generate_species_pages([animal], tmp_path)
        return (tmp_path / "species" / aid / "index.html").read_text(encoding="utf-8")

    def test_species_data_json_embedded(self, tmp_path):
        page = self._read_page(tmp_path)
        assert "const SPECIES_DATA = {" in page
        assert "const RELATED_DATA = " in page

    def test_species_data_carries_multilingual_names(self, tmp_path):
        page = self._read_page(tmp_path)
        # name.ja / name.en are required so the client script can swap titles
        # without round-tripping through animals.json.
        assert '"name":{"ja":"スズメ"' in page
        assert '"en":"Eurasian Tree Sparrow"' in page

    def test_species_data_carries_class_translations(self, tmp_path):
        page = self._read_page(tmp_path)
        assert '"class":{"ja":"鳥綱"' in page
        assert '"en":"Birds"' in page

    def test_species_data_carries_voice_method_translation(self, tmp_path):
        page = self._read_page(tmp_path)
        assert '"voiceMethod":{"ja":"鳴管","en":"Syrinx"}' in page

    def test_species_data_carries_region_translations(self, tmp_path):
        page = self._read_page(tmp_path)
        assert '"regions":[{"ja":"日本","en":"Japan"}]' in page

    def test_i18n_markers_present(self, tmp_path):
        page = self._read_page(tmp_path)
        # Every label that needs language swapping carries data-i18n so the
        # script can locate it without depending on text content.
        for marker in (
            'data-i18n="skipLink"',
            'data-i18n="topBar"',
            'data-i18n="infoSection"',
            'data-i18n="voiceMethod"',
            'data-i18n="conservation"',
            'data-i18n="regions"',
            'data-i18n="linksSection"',
            'data-i18n="shareSection"',
            'data-i18n="backLink"',
        ):
            assert marker in page, f"missing i18n marker: {marker}"

    def test_share_buttons_have_kind_attribute(self, tmp_path):
        page = self._read_page(tmp_path)
        for kind in ("x", "fb", "line", "copy"):
            assert f'data-share-kind="{kind}"' in page, f"share kind {kind} missing"
        # The copy button wraps its label so the script can rewrite text only.
        assert '<span class="copy-label">URLをコピー</span>' in page

    def test_external_links_have_kind_attribute(self, tmp_path):
        page = self._read_page(tmp_path)
        for kind in ("commons", "xc", "wiki-ja", "wiki-en"):
            assert f'data-link-kind="{kind}"' in page, f"link kind {kind} missing"

    def test_hreflang_points_to_species_page_per_locale(self, tmp_path):
        page = self._read_page(tmp_path)
        # Detail-page hreflang must point back to the species page itself per
        # locale (was previously the index modal deep link).
        for code in ("ja", "en", "ko", "zh"):
            assert (
                f'hreflang="{code}" href="{build.SITE_URL}/species/B001/?lang={code}"'
                in page
            ), f"missing hreflang for {code}"
        assert (
            f'hreflang="x-default" href="{build.SITE_URL}/species/B001/"' in page
        )

    def test_inline_script_protected_from_closing_script_tag(self, tmp_path):
        """SPECIES_DATA payload with `</script>` must not split the block.

        The embedded JSON is HTML-escaped (< > & → \\uXXXX) so a hostile name
        cannot prematurely close the inline <script> and break the page.
        """
        animal = self._make_animal("X001")
        animal["nameEN"] = 'Evil</script><script>alert(1)</script>'
        build.generate_species_pages([animal], tmp_path)
        page = (tmp_path / "species" / "X001" / "index.html").read_text(
            encoding="utf-8"
        )
        # The SPECIES_DATA assignment lives in the inline data-script block.
        # The hostile `</script>` from the input must NOT appear raw between
        # `const SPECIES_DATA` and the next `</script>` — otherwise the block
        # would be split and arbitrary script would execute.
        start = page.index("const SPECIES_DATA")
        end = page.index("</script>", start)
        body = page[start:end]
        assert "<script>alert" not in body
        # Escaped form must be present so the data round-trips faithfully.
        assert "\\u003cscript\\u003e" in body


# ── _json_for_inline_script (XSS-safe JSON emission) ────


class TestJsonForInlineScript:
    """Inline JSON emitted via _json_for_inline_script must survive being
    embedded inside a <script> block even when fields contain HTML markup.
    """

    def test_escapes_html_sensitive_characters(self):
        out = build._json_for_inline_script({"v": "</script><b>"})
        assert "</script>" not in out
        assert "\\u003c/script\\u003e" in out
        assert "\\u003cb\\u003e" in out

    def test_round_trips_unicode_payload(self):
        """Non-ASCII content stays raw (ensure_ascii=False) so JS receives the
        same string the build saw — avoids \\uXXXX bloat for CJK text."""
        out = build._json_for_inline_script({"name": "スズメ"})
        assert "スズメ" in out

    def test_produces_valid_json(self):
        import json as _json
        out = build._json_for_inline_script({"a": [1, 2, 3], "b": "x"})
        # After undoing the HTML escapes, the result must parse as JSON.
        unescaped = out.replace("\\u003c", "<").replace("\\u003e", ">").replace("\\u0026", "&")
        parsed = _json.loads(unescaped)
        assert parsed == {"a": [1, 2, 3], "b": "x"}


# ── Excel data integrity (real data) ────────────────────

class TestExcelIntegrity:
    """Assert referential integrity across Excel sheets.

    These are contract-level checks (foreign keys, required mappings) that the
    build does not enforce because it silently tolerates missing joins.
    """

    @pytest.fixture(scope="class")
    def wb(self):
        if not build.DATA_FILE.exists():
            pytest.skip("Excel data file not present")
        return openpyxl.load_workbook(build.DATA_FILE, data_only=True)

    def _collect_ids(self, ws):
        return [str(row[0]) for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]

    def test_name_mapping_matches_main_data(self, wb):
        main_ids = set(self._collect_ids(wb["メインデータ"]))
        name_ids = set(self._collect_ids(wb["名称マッピング"]))
        assert main_ids == name_ids, (
            f"name-map mismatch: missing={main_ids - name_ids}, orphan={name_ids - main_ids}"
        )

    def test_onomatopoeia_mapping_has_no_orphans(self, wb):
        # Guards against I033-style orphans: onomatopoeia entries that reference
        # a species ID not present in メインデータ. Safe to remove these entries.
        main_ids = set(self._collect_ids(wb["メインデータ"]))
        ono_ids = set(self._collect_ids(wb["オノマトペマッピング"]))
        orphans = ono_ids - main_ids
        assert not orphans, f"Orphan onomatopoeia IDs (no matching main data): {orphans}"

    def test_region_mapping_has_no_orphans(self, wb):
        main_ids = set(self._collect_ids(wb["メインデータ"]))
        reg_ids = set(self._collect_ids(wb["地域マッピング"]))
        orphans = reg_ids - main_ids
        assert not orphans, f"Orphan region mapping IDs: {orphans}"

    def test_region_mapping_uses_valid_region_ids(self, wb):
        master_rids = {str(row[0]) for row in wb["地域マスター"].iter_rows(min_row=2, values_only=True) if row[0]}
        used_rids = set()
        for row in wb["地域マッピング"].iter_rows(min_row=2, values_only=True):
            if row[0] and row[2]:
                used_rids.add(str(row[2]))
        unknown = used_rids - master_rids
        assert not unknown, f"Region IDs referenced but not in 地域マスター: {unknown}"

    def test_all_classes_have_taxonomy_entry(self, wb):
        # classEN in animals.json is populated from 分類マッピング. Missing entries
        # leave classEN empty, which silently degrades JSON consumers.
        classes = set()
        for row in wb["メインデータ"].iter_rows(min_row=2, values_only=True):
            if row[0] and row[3]:
                classes.add(str(row[3]))
        tax = {(str(row[0]), str(row[1])) for row in wb["分類マッピング"].iter_rows(min_row=2, values_only=True) if row[0] and row[1]}
        missing = [c for c in classes if ("綱", c) not in tax]
        assert not missing, f"Classes without taxonomy mapping (empty classEN): {missing}"

    def test_all_orders_have_taxonomy_entry(self, wb):
        orders = set()
        for row in wb["メインデータ"].iter_rows(min_row=2, values_only=True):
            if row[0] and row[4]:
                orders.add(str(row[4]))
        tax = {(str(row[0]), str(row[1])) for row in wb["分類マッピング"].iter_rows(min_row=2, values_only=True) if row[0] and row[1]}
        missing = [o for o in orders if ("目", o) not in tax]
        assert not missing, f"Orders without taxonomy mapping (empty orderEN): {missing}"

    def test_all_families_have_taxonomy_entry(self, wb):
        families = set()
        for row in wb["メインデータ"].iter_rows(min_row=2, values_only=True):
            if row[0] and row[5]:
                families.add(str(row[5]))
        tax = {(str(row[0]), str(row[1])) for row in wb["分類マッピング"].iter_rows(min_row=2, values_only=True) if row[0] and row[1]}
        missing = [f for f in families if ("科", f) not in tax]
        assert not missing, f"Families without taxonomy mapping (empty familyEN): {missing}"

    def test_has_voice_species_have_all_four_languages(self, wb):
        # When hasVoice=あり, we expect onomatopoeia for all 4 languages.
        main_rows = {
            str(row[0]): row
            for row in wb["メインデータ"].iter_rows(min_row=2, values_only=True)
            if row[0]
        }
        ono_by_id = {}
        for row in wb["オノマトペマッピング"].iter_rows(min_row=2, values_only=True):
            if row[0] and row[2] and row[4]:
                ono_by_id.setdefault(str(row[0]), set()).add(str(row[2]))
        expected = {"ja", "en", "ko", "zh"}
        missing = []
        for aid, row in main_rows.items():
            if row[6] == "あり":
                have = ono_by_id.get(aid, set())
                if have != expected:
                    missing.append((aid, sorted(expected - have)))
        assert not missing, f"hasVoice species missing language onomatopoeia: {missing[:5]}"

    def test_no_voice_species_have_no_onomatopoeia(self, wb):
        # Contract: hasVoice=なし implies zero onomatopoeia rows. Breaking this
        # lets bulk-generated Korean/Chinese entries leak into animals.json for
        # species that don't actually produce sound.
        main_rows = {
            str(row[0]): row
            for row in wb["メインデータ"].iter_rows(min_row=2, values_only=True)
            if row[0]
        }
        ono_by_id = {}
        for row in wb["オノマトペマッピング"].iter_rows(min_row=2, values_only=True):
            if row[0] and row[2] and row[4]:
                ono_by_id.setdefault(str(row[0]), []).append(str(row[2]))
        violations = []
        for aid, row in main_rows.items():
            if row[6] != "あり" and ono_by_id.get(aid):
                violations.append((aid, ono_by_id[aid]))
        assert not violations, (
            f"Non-voiced species have onomatopoeia entries: {violations[:5]}"
        )


# ── SW / PWA contracts ──────────────────────────────────

class TestServiceWorkerContracts:
    def test_sw_excludes_cdn_from_cache(self, tmp_path):
        # URLS list must only contain same-origin paths. Cross-origin CDN
        # (Fuse.js) is served opaque and cannot be precached deterministically.
        out = tmp_path / "sw.js"
        build.generate_sw([], out)
        sw = out.read_text(encoding="utf-8")
        assert "cdnjs" not in sw
        assert "http://" not in sw

    def test_sw_activate_clears_old_caches(self, tmp_path):
        out = tmp_path / "sw.js"
        build.generate_sw([{"id": "B001"}], out)
        sw = out.read_text(encoding="utf-8")
        # Old caches (keys !== CACHE_NAME) are deleted on activate.
        assert "caches.delete" in sw
        assert "k !== CACHE_NAME" in sw

    def test_sw_fetch_falls_back_to_root_on_navigation(self, tmp_path):
        # Offline navigation to an uncached species page falls back to /.
        out = tmp_path / "sw.js"
        build.generate_sw([], out)
        sw = out.read_text(encoding="utf-8")
        assert 'request.mode === "navigate"' in sw
        assert 'caches.match("/")' in sw

    def test_sw_parses_as_valid_javascript(self, tmp_path):
        # SW with mismatched parens silently fails to register in all browsers,
        # breaking offline caching. Use `node --check` to catch regressions.
        import shutil, subprocess
        if not shutil.which("node"):
            pytest.skip("node not available")
        out = tmp_path / "sw.js"
        build.generate_sw([{"id": "B001"}], out)
        result = subprocess.run(
            ["node", "--check", str(out)],
            capture_output=True, text=True,
        )
        assert result.returncode == 0, (
            f"Service worker has syntax error: {result.stderr}"
        )

    def test_sw_parens_are_balanced(self, tmp_path):
        # Faster paren/brace balance check that runs without node. Redundant
        # with test_sw_parses_as_valid_javascript but useful when CI lacks node.
        out = tmp_path / "sw.js"
        build.generate_sw([{"id": "B001"}], out)
        sw = out.read_text(encoding="utf-8")
        # String-literal parens must be ignored. Strip string contents first.
        import re
        stripped = re.sub(r'"[^"]*"', '""', sw)
        assert stripped.count("(") == stripped.count(")"), (
            f"Paren imbalance: ({stripped.count('(')} vs ){stripped.count(')')}"
        )
        assert stripped.count("{") == stripped.count("}"), (
            f"Brace imbalance: {{ {stripped.count('{')} vs }} {stripped.count('}')}"
        )


# ── Manifest contracts ──────────────────────────────────

class TestManifestContracts:
    def test_manifest_has_pwa_essentials(self, tmp_path):
        out = tmp_path / "manifest.json"
        build.generate_manifest(out)
        data = json.loads(out.read_text(encoding="utf-8"))
        for key in ("name", "short_name", "start_url", "display", "icons"):
            assert key in data, f"manifest missing {key}"

    def test_manifest_icons_have_type(self, tmp_path):
        out = tmp_path / "manifest.json"
        build.generate_manifest(out)
        data = json.loads(out.read_text(encoding="utf-8"))
        for icon in data["icons"]:
            assert "src" in icon and "type" in icon


# ── Sitemap date format ─────────────────────────────────

class TestSitemapDate:
    def test_sitemap_lastmod_iso_format(self, tmp_path):
        import re
        build.generate_sitemap([{"id": "B001"}], tmp_path / "sitemap.xml")
        xml = (tmp_path / "sitemap.xml").read_text(encoding="utf-8")
        dates = re.findall(r"<lastmod>([^<]+)</lastmod>", xml)
        assert dates
        for d in dates:
            # YYYY-MM-DD
            assert re.match(r"^\d{4}-\d{2}-\d{2}$", d), f"bad date format: {d}"


# ── Build_stats boundaries ──────────────────────────────

class TestBuildStatsBoundaries:
    def test_missing_onomatopoeia_key_raises(self):
        # If someone feeds animals without the onomatopoeia field, build_stats
        # should fail loudly rather than silently producing wrong counts.
        with pytest.raises((KeyError, TypeError)):
            build.build_stats([{"id": "X001"}])

    def test_lang_set_is_sorted(self):
        animals = [{"onomatopoeia": [{"lang": l} for l in ["zh", "en", "ja"]]}]
        stats = build.build_stats(animals)
        assert stats["languages"] == sorted(stats["languages"])


# ── _build_audio_ref extended boundaries ────────────────

class TestBuildAudioRefBoundaries:
    def test_bird_scientific_with_extra_whitespace(self):
        # Scientific name with leading/trailing spaces should still resolve.
        url = build._build_audio_ref("B001", " Passer montanus ", "pasmon1")
        # Implementation uses str.split() which ignores leading/trailing whitespace.
        assert "xeno-canto.org" in url and "Passer-montanus" in url

    def test_bird_three_word_scientific_name(self):
        # Trinomials (e.g., subspecies) use first two tokens.
        url = build._build_audio_ref("B001", "Passer montanus saturatus", "ignore")
        assert url == "https://xeno-canto.org/species/Passer-montanus"

    def test_no_audio_set_exact_match_required(self):
        # Substring matches must NOT trigger no-audio suppression.
        with patch.object(build, "_NO_AUDIO", {"canlup"}):
            url = build._build_audio_ref("M001", "Canis lupus", "canlup1")
            assert "macaulaylibrary.org" in url

    def test_xeno_canto_scientific_name_is_not_url_encoded(self):
        # xeno-canto accepts Genus-species in ASCII; non-ASCII scientific names
        # would require encoding. Currently NOT encoded, so document the behavior.
        # (All real scientific names are ASCII.)
        url = build._build_audio_ref("B001", "Cyanistes caeruleus", "eurbtt")
        assert url == "https://xeno-canto.org/species/Cyanistes-caeruleus"


# ── PWA icons (R3) ──────────────────────────────────────

class TestPwaIcons:
    """Raster PNG icons required by Chrome Android / Lighthouse PWA audit (M8)."""

    def test_generate_pwa_icons_produces_192_and_512(self, tmp_path):
        sizes = build.generate_pwa_icons(tmp_path)
        assert sizes == [192, 512]
        for s in sizes:
            p = tmp_path / f"icon-{s}.png"
            assert p.exists() and p.stat().st_size > 0

    def test_generate_pwa_icons_are_valid_png_with_correct_dimensions(self, tmp_path):
        # Guard against the regex change or gradient math producing corrupt
        # images that Chrome silently ignores.
        try:
            from PIL import Image
        except ImportError:
            pytest.skip("Pillow not available")
        build.generate_pwa_icons(tmp_path)
        for size in (192, 512):
            img = Image.open(tmp_path / f"icon-{size}.png")
            assert img.format == "PNG"
            assert img.size == (size, size)

    def test_generate_pwa_icons_render_silhouette(self, tmp_path):
        # Center must be distinctly lighter than corners: an all-background icon
        # means the SVG point extraction regression-returned empty.
        try:
            from PIL import Image
        except ImportError:
            pytest.skip("Pillow not available")
        build.generate_pwa_icons(tmp_path)
        img = Image.open(tmp_path / "icon-192.png")
        w, h = img.size
        center = img.getpixel((w // 2, h // 2))
        corner = img.getpixel((3, 3))
        assert sum(center[:3]) > sum(corner[:3]) + 50, (
            "PWA icon has no visible silhouette — SVG point extraction likely broken"
        )


# ── Build idempotency (R3) ──────────────────────────────

class TestBuildIdempotency:
    """Running build twice on the same day must produce byte-identical output.

    Non-determinism here would defeat SW cache-busting (cache name changes
    drive client-side cache invalidation) and waste CI cache. The only
    date-dependent field is intentional (sitemap lastmod, SW cache name).
    """

    def test_two_builds_produce_identical_dist(self):
        import hashlib, subprocess
        root = Path(__file__).resolve().parent.parent
        dist = root / "dist"

        subprocess.run(
            ["python3", "scripts/build.py"],
            cwd=root, capture_output=True, check=True,
        )
        first = {
            str(p.relative_to(dist)): hashlib.sha256(p.read_bytes()).hexdigest()
            for p in dist.rglob("*") if p.is_file()
        }
        subprocess.run(
            ["python3", "scripts/build.py"],
            cwd=root, capture_output=True, check=True,
        )
        second = {
            str(p.relative_to(dist)): hashlib.sha256(p.read_bytes()).hexdigest()
            for p in dist.rglob("*") if p.is_file()
        }
        diffs = [k for k in first if first[k] != second.get(k)]
        assert not diffs, f"Non-deterministic files: {diffs[:5]}"


# ── SW × Manifest icon coherence (R3) ───────────────────

class TestServiceWorkerManifestCoherence:
    """Cross-resource contract: PNG icons referenced by manifest should ideally
    be precached. Currently the SW URLS list omits PNG icons (M8 added them
    after the SW template was written); lazy caching still works on first
    fetch. This test documents the gap so a future change gets noticed.
    """

    def test_manifest_png_icons_lazy_cached_only(self, tmp_path):
        # If build.py is changed to precache PNG icons, update this contract.
        build.generate_manifest(tmp_path / "manifest.json", icon_sizes=(192, 512))
        build.generate_sw([], tmp_path / "sw.js")
        sw = (tmp_path / "sw.js").read_text(encoding="utf-8")
        manifest = json.loads(
            (tmp_path / "manifest.json").read_text(encoding="utf-8")
        )
        png_srcs = [i["src"] for i in manifest["icons"] if i.get("type") == "image/png"]
        assert png_srcs, "manifest has no PNG icons — M8 regression"
        # Currently PNG icons are NOT precached (lazy cache-on-fetch). Flip the
        # assertion if this design changes.
        for src in png_srcs:
            assert src not in sw, (
                f"{src} unexpectedly precached — update contract test"
            )


# ── Template SVG extraction robustness (R3) ─────────────

class TestSvgPointExtractionAnchoring:
    """Regex change: `d="..."` must only match the path data attribute, not
    attributes that happen to end in `d` (e.g., `id=`). The R2 fix anchors on
    whitespace before `d=`.
    """

    def test_does_not_match_id_attribute(self, tmp_path):
        svg = tmp_path / "t.svg"
        # `id="mydata"` ends in "ta"; path `d=` is the real target.
        svg.write_text('<svg><path id="foo" d="M 10 20 L 30 40"/></svg>')
        pts = build._parse_svg_points(svg)
        assert pts == [(10.0, 20.0), (30.0, 40.0)]

    def test_does_not_match_embedded_d_in_other_attributes(self, tmp_path):
        svg = tmp_path / "t.svg"
        # Previous naive regex `d="..."` could match inside `data-d="..."`.
        svg.write_text(
            '<svg data-d="X1 Y1"><path d="M 5 5 L 6 6"/></svg>'
        )
        pts = build._parse_svg_points(svg)
        assert (5.0, 5.0) in pts
        assert (6.0, 6.0) in pts
        # The bogus data-d coords must NOT be extracted.
        assert all(x not in {"X1", "Y1"} for pt in pts for x in pt)
