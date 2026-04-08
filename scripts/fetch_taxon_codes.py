#!/usr/bin/env python3
"""
fetch_taxon_codes.py — Macaulay Library taxonCode を学名から取得し Excel に書き込む

名称マッピングシートの学名を使い、taxonomy API で taxonCode を取得。
メインデータシートの O列 (15列目) に taxonCode を書き込む。

使い方:
    python scripts/fetch_taxon_codes.py [--dry-run]
"""

import sys
import time
import urllib.request
import urllib.parse
import json
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl required. pip install openpyxl")
    sys.exit(1)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_FILE = PROJECT_ROOT / "data" / "animal-sounds-data.xlsx"

API_URL = "https://taxonomy.api.macaulaylibrary.org/v1/taxonomy"
API_KEY = "PUB5447877383"
DELAY = 0.3  # seconds between requests

# Synonym / manual overrides: scientific name in Excel -> (query, taxonCode)
# For species where the Excel name doesn't match ML taxonomy
SYNONYMS = {
    "Dryophytes japonicus": "Hyla japonica",       # reclassified
    "Gallus gallus domesticus": "Gallus gallus",    # search parent, pick domestic
    "Canis lupus familiaris": "Canis lupus",        # search parent, pick wolf (no domestic dog entry)
    "Capra aegagrus hircus": "Capra hircus",        # domestic goat
    "Hypotaenidia okinawae": "Gallirallus okinawae",  # reclassified
    "Bufo japonicus": "Bufo japonicus formosus",    # try subspecies
}

# Species with no ML entry at all — hardcode taxonCode or leave empty
MANUAL_CODES = {
    "Canis lupus familiaris": "",     # domestic dog: no separate ML entry
    "Gallus gallus domesticus": "redjun1",  # Red Junglefowl (Domestic type)
    "Bufo japonicus": "",             # no ML entry
    "Mecopoda nipponensis": "",
    "Conocephalus melas": "",
    "Platypleura kaempferi": "",
    "Terpnosia vacua": "",
    "Lyristes japonicus": "",
    "Cybister japonicus": "",
    "Cynops orientalis": "",
    "Tylototriton asperatus": "",
}


def fetch_taxon_code(scientific_name):
    """Query ML taxonomy API, return (code, matched_name) or (None, None)."""
    params = urllib.parse.urlencode({"q": scientific_name, "key": API_KEY})
    url = f"{API_URL}?{params}"
    req = urllib.request.Request(url, headers={"User-Agent": "koe-zukan-build/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        print(f"  ERROR fetching {scientific_name}: {e}")
        return None, None

    if not data:
        return None, None

    # Prefer exact species match (scientific name appears after " - ")
    for item in data:
        name_part = item.get("name", "")
        if f" - {scientific_name}" in name_part:
            return item["code"], name_part
    # Fallback to first result
    return data[0]["code"], data[0].get("name", "")


def main():
    dry_run = "--dry-run" in sys.argv

    wb = openpyxl.load_workbook(DATA_FILE)

    # Build scientific name map from naming sheet
    ws_name = wb["名称マッピング"]
    sci_map = {}
    for row in ws_name.iter_rows(min_row=2, values_only=True):
        if row[0] and row[2]:
            sci_map[str(row[0])] = str(row[2])

    # Write header in O1 if missing
    ws_main = wb["メインデータ"]
    if ws_main.cell(1, 15).value != "taxonCode":
        ws_main.cell(1, 15, "taxonCode")

    total = 0
    found = 0
    skipped = 0
    not_found = []

    for r, row in enumerate(ws_main.iter_rows(min_row=2), start=2):
        aid = row[0].value
        if not aid:
            continue
        aid = str(aid)

        # Skip if already has taxonCode
        existing = ws_main.cell(r, 15).value
        if existing:
            skipped += 1
            continue

        sci = sci_map.get(aid)
        if not sci:
            continue

        total += 1

        # Check manual overrides first
        if sci in MANUAL_CODES:
            code = MANUAL_CODES[sci]
            if code:
                found += 1
                print(f"  {aid} {sci} -> {code} (manual)")
                if not dry_run:
                    ws_main.cell(r, 15, code)
            else:
                not_found.append((aid, sci))
                print(f"  {aid} {sci} -> SKIPPED (no ML entry)")
            time.sleep(DELAY)
            continue

        # Try synonym if available
        query = SYNONYMS.get(sci, sci)
        code, matched = fetch_taxon_code(query)
        if code:
            found += 1
            print(f"  {aid} {sci} -> {code} ({matched})")
            if not dry_run:
                ws_main.cell(r, 15, code)
        else:
            not_found.append((aid, sci))
            print(f"  {aid} {sci} -> NOT FOUND")

        time.sleep(DELAY)

    print()
    print(f"Total queried: {total}, Found: {found}, Skipped (existing): {skipped}")
    if not_found:
        print(f"Not found ({len(not_found)}):")
        for aid, sci in not_found:
            print(f"  {aid} {sci}")

    if not dry_run and found > 0:
        wb.save(DATA_FILE)
        print(f"\nSaved to {DATA_FILE}")
    elif dry_run:
        print("\n(dry-run, no changes saved)")


if __name__ == "__main__":
    main()
