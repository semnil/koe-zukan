#!/usr/bin/env python3
"""
build.py — Excel → 静的サイト生成スクリプト

data/animal-sounds-data.xlsx を読み込み、dist/ に静的サイトを出力する。
依存: openpyxl, Pillow (pip install openpyxl Pillow)

使い方:
    python scripts/build.py
"""

import html as html_mod
import json
import os
import shutil
import sys
from datetime import date
from pathlib import Path
from urllib.parse import quote, urlparse

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_FILE = PROJECT_ROOT / "data" / "animal-sounds-data.xlsx"
DIST_DIR = PROJECT_ROOT / "dist"
TEMPLATE_FILE = PROJECT_ROOT / "templates" / "index.html"
ASSETS_DIR = PROJECT_ROOT / "assets"
NO_AUDIO_FILE = PROJECT_ROOT / "data" / "no-audio.json"
SITE_URL = "https://koe-zukan.semnil.com"


def _load_no_audio():
    """Load taxonCodes with no audio on Macaulay Library."""
    if NO_AUDIO_FILE.exists():
        with open(NO_AUDIO_FILE, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


_NO_AUDIO = _load_no_audio()


def _build_audio_ref(aid, scientific_name, taxon_code):
    """Build audio reference URL from taxonCode.

    Birds use xeno-canto (by scientific name).
    Others use Macaulay Library (by taxonCode), skipped if no audio.
    """
    if not scientific_name:
        return ""
    if aid.startswith("B"):
        # xeno-canto: Genus-species format
        parts = scientific_name.split()
        if len(parts) >= 2:
            return f"https://xeno-canto.org/species/{parts[0]}-{parts[1]}"
        return ""
    if taxon_code and taxon_code not in _NO_AUDIO:
        return f"https://search.macaulaylibrary.org/catalog?taxonCode={taxon_code}&mediaType=audio"
    return ""


def extract_data(wb):
    """Excel workbook → Python dicts"""

    # ── 名称マッピング ──
    ws_name = wb["名称マッピング"]
    name_map = {}
    for row in ws_name.iter_rows(min_row=2, values_only=True):
        if row[0]:
            name_map[str(row[0])] = {
                "scientificName": row[2] or "",
                "englishName": row[3] or "",
                "altJA": row[4] or "",
                "altEN": row[5] or "",
            }

    # ── オノマトペマッピング ──
    ws_ono = wb["オノマトペマッピング"]
    ono_map = {}
    for row in ws_ono.iter_rows(min_row=2, values_only=True):
        if row[0]:
            aid = str(row[0])
            if aid not in ono_map:
                ono_map[aid] = []
            ono_map[aid].append({
                "lang": row[2] or "",
                "langName": row[3] or "",
                "onomatopoeia": row[4] or "",
                "scene": row[5] or "",
                "note": row[6] or "",
            })

    # ── 地域マスター ──
    ws_rm = wb["地域マスター"]
    region_master = {}
    for row in ws_rm.iter_rows(min_row=2, values_only=True):
        if row[0]:
            region_master[str(row[0])] = {
                "nameJA": row[1] or "",
                "nameEN": row[2] or "",
                "areaCode": row[3] or "",
            }

    # ── 地域マッピング ──
    ws_reg = wb["地域マッピング"]
    region_map = {}
    for row in ws_reg.iter_rows(min_row=2, values_only=True):
        if row[0]:
            aid = str(row[0])
            rid = str(row[2]) if row[2] else ""
            if aid not in region_map:
                region_map[aid] = []
            if rid:
                region_map[aid].append(rid)

    # ── 分類マッピング ──
    ws_tax = wb["分類マッピング"]
    tax_map = {}
    for row in ws_tax.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            tax_map[(str(row[0]), str(row[1]))] = {
                "scientificName": row[2] or "",
                "englishName": row[3] or "",
            }

    # ── メインデータ → animals ──
    ws = wb["メインデータ"]
    animals = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        aid = str(row[0])
        names = name_map.get(aid, {})
        onos = ono_map.get(aid, [])
        regions = region_map.get(aid, [])

        resolved_regions = []
        for rid in regions:
            rm = region_master.get(rid, {})
            resolved_regions.append({
                "id": rid,
                "nameJA": rm.get("nameJA", ""),
                "nameEN": rm.get("nameEN", ""),
            })

        class_en = tax_map.get(("綱", str(row[3])), {}).get("englishName", "") if row[3] else ""
        order_en = tax_map.get(("目", str(row[4])), {}).get("englishName", "") if row[4] else ""
        family_en = tax_map.get(("科", str(row[5])), {}).get("englishName", "") if row[5] else ""

        animal = {
            "id": aid,
            "nameJA": row[1] or "",
            "scientificName": names.get("scientificName", ""),
            "nameEN": names.get("englishName", ""),
            "altJA": names.get("altJA", ""),
            "altEN": names.get("altEN", ""),
            "phylum": row[2] or "",
            "class": row[3] or "",
            "classEN": class_en,
            "order": row[4] or "",
            "orderEN": order_en,
            "family": row[5] or "",
            "familyEN": family_en,
            "hasVoice": row[6] or "",
            "onomatopoeiaJA": row[7] or "",
            "voiceMethod": row[8] or "",
            "habitat": row[9] or "",
            "conservation": row[10] or "",
            "imageRef": row[11] or "",
            "note": row[12] or "",
            "audioRef": _build_audio_ref(aid, names.get("scientificName", ""),
                                         str(row[14]) if len(row) > 14 and row[14] else ""),
            "onomatopoeia": onos,
            "regions": resolved_regions,
        }
        animals.append(animal)

    # ── 地域マスター (flat) ──
    regions_data = []
    for rid, rm in sorted(region_master.items()):
        regions_data.append({"id": rid, **rm})

    return animals, regions_data


def build_stats(animals):
    """Generate stats for template injection."""
    total = len(animals)
    ono_count = sum(len(a["onomatopoeia"]) for a in animals)
    langs = set()
    for a in animals:
        for o in a["onomatopoeia"]:
            if o["lang"]:
                langs.add(o["lang"])
    return {
        "total_species": total,
        "total_onomatopoeia": ono_count,
        "languages": sorted(langs),
        "language_count": len(langs),
    }


def write_json(data, filepath):
    """Write JSON with UTF-8, no ASCII escaping."""
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    size = os.path.getsize(filepath)
    print(f"  {filepath.name}: {size:,} bytes")


def _apply_placeholders(html, mapping):
    """Replace {{KEY}} placeholders in HTML with values from mapping dict."""
    import re
    for key, value in mapping.items():
        html = html.replace(f"{{{{{key}}}}}", str(value))
    remaining = re.findall(r"\{\{[A-Z_]+\}\}", html)
    if remaining:
        print(f"  WARNING: unreplaced placeholders: {remaining}")
    return html


def generate_html(animals, template_path, output_path):
    """Read HTML template and inject data if template exists."""
    if template_path.exists():
        with open(template_path, "r", encoding="utf-8") as f:
            html = f.read()
    else:
        # Fallback: copy from dist if template doesn't exist yet
        fallback = output_path
        if fallback.exists():
            print(f"  Using existing {fallback.name} (no template found)")
            return
        print(f"  WARNING: No template at {template_path} and no existing {output_path.name}")
        return

    stats = build_stats(animals)
    html = _apply_placeholders(html, {
        "SITE_URL": SITE_URL,
        "SPECIES_COUNT": stats['total_species'],
        "LANGUAGE_COUNT": stats['language_count'],
        "ONOMATOPOEIA_COUNT": stats['total_onomatopoeia'],
    })

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    size = os.path.getsize(output_path)
    print(f"  {output_path.name}: {size:,} bytes")


def _parse_svg_points(svg_path):
    """Extract polygon points from favicon SVG path data."""
    import re
    with open(svg_path, "r", encoding="utf-8") as f:
        svg = f.read()
    match = re.search(r'd="([^"]+)"', svg)
    if not match:
        return []
    coords = re.findall(r"[\d.]+", match.group(1))
    return [(float(coords[i]), float(coords[i + 1])) for i in range(0, len(coords), 2)]


def _find_cjk_fonts():
    """Find CJK fonts across platforms. Returns dict with 'ja', 'ko', 'zh' keys."""
    import sys as _sys
    if _sys.platform == "win32":
        ja_candidates = ["C:/Windows/Fonts/meiryo.ttc", "C:/Windows/Fonts/msgothic.ttc"]
        ko_candidates = ["C:/Windows/Fonts/malgun.ttf", "C:/Windows/Fonts/malgunbd.ttf"]
        zh_candidates = ["C:/Windows/Fonts/msyh.ttc", "C:/Windows/Fonts/simsun.ttc"]
    else:
        # Linux: Noto Sans CJK covers all
        noto = ["/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
                "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
                "/usr/share/fonts/noto-cjk/NotoSansCJK-Regular.ttc"]
        ja_candidates = ko_candidates = zh_candidates = noto

    def _find(candidates):
        for p in candidates:
            if Path(p).exists():
                return p
        return None

    ja = _find(ja_candidates)
    return {"ja": ja, "ko": _find(ko_candidates) or ja, "zh": _find(zh_candidates) or ja}


def _ogp_base_image():
    """Create base OGP image (gradient + cat silhouette). Returns (Image, font_path) or None."""
    try:
        from PIL import Image, ImageDraw
    except ImportError:
        return None

    W, H = 1200, 630
    img = Image.new("RGB", (W, H), (42, 38, 52))
    draw = ImageDraw.Draw(img)

    for y in range(H):
        r = int(42 + (58 - 42) * y / H)
        g = int(38 + (48 - 38) * y / H)
        b = int(52 + (68 - 52) * y / H)
        draw.line([(0, y), (W, y)], fill=(r, g, b))

    favicon = ASSETS_DIR / "favicon.svg"
    raw_points = _parse_svg_points(favicon) if favicon.exists() else []
    if raw_points:
        cx, cy, cat_size = 200, 315, 280
        scale = cat_size / 64.0
        points = [(cx + (x - 32) * scale, cy + (y - 38) * scale) for x, y in raw_points]
        draw.polygon(points, fill=(91, 74, 122))

    return img


def generate_ogp(stats, output_path, base_image=None):
    """Generate top-page OGP image (1200x630) with dynamic stats."""
    from PIL import ImageDraw, ImageFont

    base = base_image or _ogp_base_image()
    if base is None:
        print("  WARNING: Pillow not installed, skipping OGP image generation")
        return

    img = base.copy()
    draw = ImageDraw.Draw(img)

    fonts = _find_cjk_fonts()
    font_path = fonts["ja"]
    if font_path:
        title_font = ImageFont.truetype(font_path, 52)
        sub_font = ImageFont.truetype(font_path, 24)
        en_font = ImageFont.truetype(font_path, 20)
    else:
        title_font = ImageFont.load_default()
        sub_font = ImageFont.load_default()
        en_font = ImageFont.load_default()

    white = (255, 255, 255)
    light = (200, 195, 210)
    draw.text((440, 200), "動物の鳴き声図鑑", fill=white, font=title_font)
    draw.text((440, 280), "Animal Sound Encyclopedia", fill=light, font=en_font)
    draw.text((440, 340),
              f"{stats['total_species']}種 × {stats['language_count']}言語のオノマトペを収録",
              fill=light, font=sub_font)
    draw.rectangle([(440, 380), (750, 383)], fill=(5, 150, 105))

    img.save(output_path, "PNG", optimize=True)
    size = os.path.getsize(output_path)
    print(f"  {output_path.name}: {size:,} bytes")


def generate_species_ogp(animals, dist_dir, base_image=None):
    """Generate per-species OGP images at /species/{id}/ogp.png."""
    from PIL import ImageDraw, ImageFont

    base = base_image or _ogp_base_image()
    if base is None:
        print("  WARNING: Pillow not installed, skipping species OGP")
        return

    fonts = _find_cjk_fonts()
    ja_path = fonts["ja"]
    if ja_path:
        name_font = ImageFont.truetype(ja_path, 48)
        ono_font = ImageFont.truetype(ja_path, 64)
        small_font = ImageFont.truetype(ja_path, 18)
    else:
        name_font = ImageFont.load_default()
        ono_font = ImageFont.load_default()
        small_font = ImageFont.load_default()

    # Per-language sub fonts for onomatopoeia
    lang_fonts = {}
    for lang, fp in fonts.items():
        lang_fonts[lang] = ImageFont.truetype(fp, 22) if fp else ImageFont.load_default()
    # English uses ja font
    lang_fonts["en"] = lang_fonts["ja"]

    white = (255, 255, 255)
    light = (200, 195, 210)
    accent = (5, 150, 105)
    count = 0

    for a in animals:
        img = base.copy()
        draw = ImageDraw.Draw(img)

        # Species name (JA)
        draw.text((440, 160), a["nameJA"], fill=white, font=name_font)

        # English name + scientific name
        en_line = a.get("nameEN", "")
        if a.get("scientificName"):
            en_line += f"  ({a['scientificName']})" if en_line else a["scientificName"]
        draw.text((440, 225), en_line, fill=light, font=small_font)

        # Accent line
        draw.rectangle([(440, 260), (750, 263)], fill=accent)

        # Onomatopoeia (JA) — large
        ja_ono = a.get("onomatopoeiaJA", "")
        if ja_ono:
            draw.text((440, 280), ja_ono, fill=accent, font=ono_font)

        # Other language onomatopoeia (use per-language font)
        other_onos = []
        for o in a.get("onomatopoeia", []):
            if o["lang"] != "ja" and o["onomatopoeia"]:
                label = LANG_LABELS.get(o["lang"], o["lang"])
                other_onos.append((o["lang"], f"{label}: {o['onomatopoeia']}"))
        if other_onos:
            y_pos = 360
            for lang, line in other_onos[:3]:
                font = lang_fonts.get(lang, lang_fonts["ja"])
                draw.text((440, y_pos), line, fill=light, font=font)
                y_pos += 30

        # Site name
        draw.text((440, 470), "動物の鳴き声図鑑", fill=(150, 145, 160), font=small_font)

        out_dir = dist_dir / "species" / a["id"]
        out_dir.mkdir(parents=True, exist_ok=True)
        img.save(out_dir / "ogp.png", "PNG", optimize=True)
        count += 1

    total_size = sum(
        f.stat().st_size for f in (dist_dir / "species").rglob("ogp.png")
    )
    print(f"  {count} species OGP images ({total_size:,} bytes total)")


SPECIES_TEMPLATE = PROJECT_ROOT / "templates" / "species.html"

CONSERVATION_JA = {
    "LC": "低懸念", "NT": "準絶滅危惧", "VU": "危急", "EN": "絶滅危惧",
    "CR": "深刻な危機", "DD": "データ不足", "NE": "未評価", "EW": "野生絶滅", "EX": "絶滅",
}

LANG_LABELS = {"ja": "日本語", "en": "English", "ko": "한국어", "zh": "中文"}


def generate_species_pages(animals, dist_dir):
    """Generate individual HTML pages for each species at /species/{id}/index.html."""
    if not SPECIES_TEMPLATE.exists():
        print("  WARNING: species.html template not found, skipping")
        return 0

    with open(SPECIES_TEMPLATE, "r", encoding="utf-8") as f:
        template = f.read()

    esc = html_mod.escape
    count = 0
    for a in animals:
        aid = a["id"]

        # Onomatopoeia by language
        ono_html_parts = []
        for o in a["onomatopoeia"]:
            if o["onomatopoeia"]:
                lang_label = esc(LANG_LABELS.get(o["lang"], o["lang"]))
                ono_html_parts.append(
                    f'<div class="ono-cell">'
                    f'<div class="ono-cell-lang">{lang_label}</div>'
                    f'<div class="ono-cell-text">{esc(o["onomatopoeia"])}</div>'
                    f'{"<div class=\"ono-cell-scene\">" + esc(o["scene"]) + "</div>" if o["scene"] else ""}'
                    f'</div>'
                )
        ono_section = ""
        if ono_html_parts:
            ono_section = (
                '<div class="detail-section"><h3>オノマトペ / Onomatopoeia</h3>'
                f'<div class="ono-grid">{"".join(ono_html_parts)}</div></div>'
            )

        # Conservation
        cons = a.get("conservation", "")
        cons_label = CONSERVATION_JA.get(cons, "")
        cons_display = f"{esc(cons)} ({cons_label})" if cons_label else esc(cons)

        # Regions
        regions_text = "、".join(esc(r["nameJA"]) for r in a["regions"]) or esc(a.get("habitat", ""))

        # External links
        links = []
        if a.get("imageRef"):
            links.append(f'<a href="{esc(a["imageRef"])}" target="_blank" rel="noopener">📷 Wikimedia Commons</a>')
        if a.get("audioRef"):
            audio_label = "xeno-canto" if aid.startswith(("A", "B")) else "Macaulay Library"
            links.append(f'<a href="{esc(a["audioRef"])}" target="_blank" rel="noopener">🔊 {audio_label}</a>')
        links.append(f'<a href="https://ja.wikipedia.org/wiki/{quote(a["nameJA"], safe="")}" target="_blank" rel="noopener">📖 Wikipedia (JA)</a>')
        if a.get("nameEN"):
            links.append(f'<a href="https://en.wikipedia.org/wiki/{quote(a["nameEN"], safe="")}" target="_blank" rel="noopener">📖 Wikipedia (EN)</a>')
        links_html = "\n".join(links)

        # Description for meta
        ja_ono = a.get("onomatopoeiaJA", "")
        desc = f"{a['nameJA']}（{a.get('nameEN', '')}）の鳴き声"
        if ja_ono:
            desc += f"「{ja_ono}」"
        desc += "。多言語オノマトペと基本情報"

        alt_en = a.get("altEN", "")
        note = a.get("note", "")
        note_html = f'<div class="detail-row"><span class="detail-label">備考</span><span>{esc(note)}</span></div>' if note else ""
        page = _apply_placeholders(template, {
            "SITE_URL": SITE_URL,
            "ID": esc(aid),
            "NAME_JA": esc(a["nameJA"]),
            "NAME_EN": esc(a.get("nameEN", "")),
            "SCIENTIFIC_NAME": esc(a.get("scientificName", "")),
            "ALT_EN": f" ({esc(alt_en)})" if alt_en else "",
            "CLASS": esc(a.get("class", "")),
            "ORDER": esc(a.get("order", "")),
            "FAMILY": esc(a.get("family", "")),
            "VOICE_METHOD": esc(a.get("voiceMethod", "") or "—"),
            "CONSERVATION": cons_display,
            "REGIONS": regions_text,
            "NOTE": note_html,
            "DESCRIPTION": esc(desc),
            "ONO_SECTION": ono_section,
            "LINKS": links_html,
        })

        # Share buttons
        share_url = f"{SITE_URL}/species/{aid}/"
        share_text = f"{a['nameJA']}（{a.get('nameEN', '')}）の鳴き声"
        eu = quote(share_url, safe="")
        et = quote(share_text, safe="")
        share_html = (
            f'<a class="share-btn" href="https://twitter.com/intent/tweet?text={et}&url={eu}" target="_blank" rel="noopener">\U0001d54f Post</a>'
            f'<a class="share-btn" href="https://www.facebook.com/sharer/sharer.php?u={eu}" target="_blank" rel="noopener">Facebook</a>'
            f'<a class="share-btn" href="https://social-plugins.line.me/lineit/share?url={eu}" target="_blank" rel="noopener">LINE</a>'
            f'<button class="share-btn" onclick="copyShareUrl(\'{esc(share_url)}\', this)">\U0001f4cb URL\u3092\u30b3\u30d4\u30fc</button>'
        )
        page = page.replace("{{SHARE_BUTTONS}}", share_html)

        out_dir = dist_dir / "species" / aid
        out_dir.mkdir(parents=True, exist_ok=True)
        with open(out_dir / "index.html", "w", encoding="utf-8") as f:
            f.write(page)
        count += 1

    total_size = sum(
        f.stat().st_size for f in (dist_dir / "species").rglob("index.html")
    )
    print(f"  {count} species pages ({total_size:,} bytes total)")
    return count


def generate_sitemap(animals, output_path):
    """Generate sitemap.xml with top page and deep links for each species."""
    today = date.today().isoformat()
    urls = [f'  <url><loc>{SITE_URL}/</loc><lastmod>{today}</lastmod><priority>1.0</priority></url>']
    for a in animals:
        urls.append(f'  <url><loc>{SITE_URL}/species/{a["id"]}/</loc><lastmod>{today}</lastmod><priority>0.6</priority></url>')
    xml = ('<?xml version="1.0" encoding="UTF-8"?>\n'
           '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
           + "\n".join(urls) + "\n"
           '</urlset>\n')
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(xml)
    size = os.path.getsize(output_path)
    print(f"  {output_path.name}: {size:,} bytes ({len(urls)} URLs)")


def generate_manifest(output_path):
    """Generate PWA manifest.json."""
    manifest = {
        "name": "動物の鳴き声図鑑",
        "short_name": "鳴き声図鑑",
        "description": "多言語対応の動物オノマトペ検索サイト",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#fafaf9",
        "theme_color": "#059669",
        "icons": [
            {"src": "/favicon.svg", "type": "image/svg+xml", "sizes": "any"}
        ],
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    print(f"  {output_path.name}: {os.path.getsize(output_path):,} bytes")


def generate_sw(animals, output_path):
    """Generate service worker for offline caching."""
    # Cache version based on build date + species count for busting
    version = f"v{date.today().isoformat()}-{len(animals)}"
    urls_to_cache = [
        "/",
        "/animals.json",
        "/regions.json",
        "/favicon.svg",
        "/manifest.json",
    ]
    sw = f"""const CACHE_NAME = "koe-zukan-{version}";
const URLS = {json.dumps(urls_to_cache)};

self.addEventListener("install", e => {{
  e.waitUntil(caches.open(CACHE_NAME).then(c => c.addAll(URLS)));
  self.skipWaiting();
}});

self.addEventListener("activate", e => {{
  e.waitUntil(
    caches.keys().then(keys =>
      Promise.all(keys.filter(k => k !== CACHE_NAME).map(k => caches.delete(k)))
    )
  );
  self.clients.claim();
}});

self.addEventListener("fetch", e => {{
  e.respondWith(
    caches.match(e.request).then(r => r || fetch(e.request).then(resp => {{
      if (resp.ok && e.request.method === "GET") {{
        const clone = resp.clone();
        caches.open(CACHE_NAME).then(c => c.put(e.request, clone));
      }}
      return resp;
    }})).catch(() => caches.match("/"))
  );
}});
"""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(sw)
    print(f"  {output_path.name}: {os.path.getsize(output_path):,} bytes")


def main():
    print(f"=== koe-zukan build ===")
    print(f"Data source: {DATA_FILE}")
    print(f"Output dir:  {DIST_DIR}")
    print()

    if not DATA_FILE.exists():
        print(f"Error: Data file not found: {DATA_FILE}")
        sys.exit(1)

    # Load Excel
    print("Loading Excel...")
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    # Extract data
    print("Extracting data...")
    animals, regions = extract_data(wb)
    stats = build_stats(animals)
    print(f"  Species: {stats['total_species']}")
    print(f"  Onomatopoeia: {stats['total_onomatopoeia']}")
    print(f"  Languages: {', '.join(stats['languages'])}")

    # Ensure dist directory
    DIST_DIR.mkdir(parents=True, exist_ok=True)

    # Write JSON
    print("Writing JSON...")
    write_json(animals, DIST_DIR / "animals.json")
    write_json(regions, DIST_DIR / "regions.json")

    # Generate HTML
    print("Generating HTML...")
    generate_html(animals, TEMPLATE_FILE, DIST_DIR / "index.html")

    # Generate sitemap
    print("Generating sitemap...")
    generate_sitemap(animals, DIST_DIR / "sitemap.xml")

    # CNAME for GitHub Pages custom domain
    cname_domain = urlparse(SITE_URL).netloc
    with open(DIST_DIR / "CNAME", "w") as f:
        f.write(cname_domain + "\n")
    print(f"  CNAME: {cname_domain}")

    # Generate species pages
    print("Generating species pages...")
    generate_species_pages(animals, DIST_DIR)

    # Generate OGP images
    print("Generating OGP images...")
    ogp_base = _ogp_base_image()
    generate_ogp(stats, DIST_DIR / "ogp.png", ogp_base)
    generate_species_ogp(animals, DIST_DIR, ogp_base)

    # Generate PWA files
    print("Generating PWA files...")
    generate_manifest(DIST_DIR / "manifest.json")
    generate_sw(animals, DIST_DIR / "sw.js")

    # Copy assets
    if ASSETS_DIR.exists():
        print("Copying assets...")
        for src in ASSETS_DIR.iterdir():
            if src.is_file():
                dst = DIST_DIR / src.name
                shutil.copy2(src, dst)
                print(f"  {src.name}: {os.path.getsize(dst):,} bytes")

    print()
    print(f"Build complete! {stats['total_species']} species → {DIST_DIR}/")
    print(f"Deploy the contents of {DIST_DIR}/ to Cloudflare Pages.")


if __name__ == "__main__":
    main()
